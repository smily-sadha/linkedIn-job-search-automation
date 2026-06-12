"""Weekly summary report — text file on disk + optional Telegram push."""
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from config.config import EXCEL_FILE, ROOT
from tracker.excel_tracker import (
    S_APPLIED, S_COLD, S_MANUAL, S_SKIPPED, S_WALKIN, init_excel,
)
from utils.logger import get_logger

logger = get_logger("weekly_report")
_REPORT_DIR = ROOT / "reports"


def _rows_in_last_days(ws, date_col: int, days: int) -> int:
    cutoff = date.today() - timedelta(days=days)
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[date_col - 1] if len(row) >= date_col else None
        try:
            if val and date.fromisoformat(str(val)[:10]) >= cutoff:
                count += 1
        except (ValueError, TypeError):
            continue
    return count


def generate_weekly_report() -> str:
    init_excel()
    wb = load_workbook(EXCEL_FILE, read_only=True)

    applied = wb[S_APPLIED]
    this_week = _rows_in_last_days(applied, 4, 7)
    last_week = _rows_in_last_days(applied, 4, 14) - this_week

    # Best source this week (by applied rows).
    sources = [r[4] for r in applied.iter_rows(min_row=2, values_only=True) if len(r) >= 5 and r[4]]
    best_source = max(set(sources), key=sources.count) if sources else "-"

    walkins = wb[S_WALKIN].max_row - 1
    cold = wb[S_COLD].max_row - 1
    manual = wb[S_MANUAL].max_row - 1
    skipped = wb[S_SKIPPED].max_row - 1
    wb.close()

    delta = this_week - last_week
    trend = "up" if delta > 0 else ("down" if delta < 0 else "flat")

    report = (
        f"Weekly Job-Hunt Report - {datetime.now():%d %b %Y}\n"
        f"{'=' * 40}\n"
        f"Applied this week:   {this_week}\n"
        f"Applied last week:   {last_week}  (trend: {trend} {delta:+d})\n"
        f"Best source:         {best_source}\n"
        f"Walk-ins logged:     {walkins}\n"
        f"Cold mails sent:     {cold}\n"
        f"Manual queue:        {manual}\n"
        f"Skipped (total):     {skipped}\n"
    )

    _REPORT_DIR.mkdir(parents=True, exist_ok=True)
    out = _REPORT_DIR / f"weekly_{date.today().isoformat()}.txt"
    out.write_text(report, encoding="utf-8")
    logger.info("Weekly report written: %s", out)

    try:
        from telegram_bot.notifier import send_report
        send_report(report)
    except Exception as exc:
        logger.info("Weekly report Telegram push skipped: %s", exc)

    return report
