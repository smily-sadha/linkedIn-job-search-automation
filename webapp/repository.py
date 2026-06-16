"""Read/write access to the Excel workbook for the web UI.

Everything here is a thin wrapper over the existing tracker so the web layer
never re-implements business logic. Reads use read_only workbooks; the few
writes reuse the same openpyxl path the rest of the app uses.

Note (Windows): openpyxl cannot write while the workbook is open in Excel.
Write helpers raise WorkbookLocked so the UI can show a friendly message.
"""
import re
from datetime import date, datetime

from openpyxl import load_workbook

from config.config import EXCEL_FILE
from mailers import cold_mailer
from tracker.excel_tracker import (
    HEADERS, S_APPLIED, S_MANUAL, S_WALKIN, WB_LOCK, init_excel, log_applied,
    refresh_dashboard, save_workbook,
)
from tracker.status_manager import get_pipeline_summary, update_status
from utils.logger import get_logger

logger = get_logger("webapp.repository")


class WorkbookLocked(Exception):
    """Raised when the workbook can't be written (usually open in Excel)."""


def _rows_as_dicts(sheet: str) -> list[dict]:
    """Return every non-empty data row of `sheet` as a header-keyed dict."""
    init_excel()
    headers = HEADERS[sheet]
    # Serialise with writers (background runs) so we never read a half-saved file.
    with WB_LOCK:
        wb = load_workbook(EXCEL_FILE, read_only=True)
        ws = wb[sheet]
        out: list[dict] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            out.append({h: row[i] if i < len(row) else None for i, h in enumerate(headers)})
        wb.close()
        return out


# ── Resume-match model ──────────────────────────────────────────────────────
# Map the 1-10 fit score to a "Match" percentage (how well you fit the role),
# plus a display tier. This is an honest resume-to-job fit signal for deciding
# what to apply to first — NOT a real probability of being hired. Tunable here.
_PCT_BY_SCORE = {
    10: 92, 9: 85, 8: 76, 7: 66, 6: 55,
    5: 43, 4: 32, 3: 22, 2: 13, 1: 6, 0: 4,
}


def selection_probability(score) -> dict:
    """Return {pct, tier, label, score} describing how well you match the role."""
    try:
        n = int(score)
    except (TypeError, ValueError):
        n = 0
    n = max(0, min(10, n))
    pct = _PCT_BY_SCORE.get(n, 4)
    if pct >= 65:
        tier, label = "high", "Strong match"
    elif pct >= 40:
        tier, label = "mid", "Moderate match"
    else:
        tier, label = "low", "Weak match"
    return {"pct": pct, "tier": tier, "label": label, "score": n}


# ── Application-window model ────────────────────────────────────────────────--
# Most online postings stay open ~30 days. We don't get a hard deadline from the
# public sources, so this is an estimate from the posting date (or, if unknown,
# the date we first found it). It exists to stop you wasting time applying to
# roles that have almost certainly closed.
_APPLY_WINDOW_DAYS = 30


def _parse_date(value) -> "date | None":
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def application_window(posted, date_found) -> dict:
    """Estimate whether a posting's application window is still open.

    Returns {tier, label, days_left, open}. tier is one of
    open | soon | closed | unknown.
    """
    ref = _parse_date(posted) or _parse_date(date_found)
    if ref is None:
        return {"tier": "unknown", "label": "Date unknown", "days_left": None, "open": True}
    days_left = _APPLY_WINDOW_DAYS - (date.today() - ref).days
    if days_left <= 0:
        return {"tier": "closed", "label": "Likely closed", "days_left": 0, "open": False}
    if days_left <= 7:
        return {"tier": "soon", "label": f"~{days_left}d left", "days_left": days_left, "open": True}
    return {"tier": "open", "label": f"~{days_left}d left", "days_left": days_left, "open": True}


# ── Posting-freshness model ──────────────────────────────────────────────────
# Sources hand us the "Posted" value in two shapes: a real date (YYYY-MM-DD from
# the API sources) or relative text ("3 days ago", "30+ days ago", "5 hours ago"
# from LinkedIn Gmail alerts). We normalise both into an age-in-days so the UI
# can carve out a "Fresh" space for just-posted roles instead of one flat list.
_FRESH_MAX_DAYS = 3  # posted within this many days counts as "fresh"
_REL_AGE_RE = re.compile(r"(\d+)\s*\+?\s*(hour|day|week|month)", re.I)


def _posted_age_days(posted, date_found) -> "int | None":
    """Best-effort age in days since a job was posted.

    Parses real dates and the relative text some sources give ("today",
    "yesterday", "3 days ago", "30+ days ago", "5 hours ago"). Falls back to
    Date Found when Posted is missing/unparseable. None when truly unknown.
    """
    d = _parse_date(posted)
    if d is not None:
        return max(0, (date.today() - d).days)
    text = str(posted or "").strip().lower()
    if text:
        if any(w in text for w in ("just now", "today", "hour", "minute")):
            return 0
        if "yesterday" in text:
            return 1
        m = _REL_AGE_RE.search(text)
        if m:
            n, unit = int(m.group(1)), m.group(2).lower()
            days = {"hour": 0, "day": n, "week": n * 7, "month": n * 30}[unit]
            # "30+ days ago" means at least n — push past the bucket edge.
            return days + 1 if "+" in text else days
    d = _parse_date(date_found)
    if d is not None:
        return max(0, (date.today() - d).days)
    return None


def freshness(posted, date_found) -> dict:
    """Return {tier, label, fresh, days} describing how recently a role posted.

    tier is one of fresh | recent | old | unknown. `fresh` is True for postings
    within _FRESH_MAX_DAYS so the UI can filter the dedicated Fresh space.
    """
    days = _posted_age_days(posted, date_found)
    if days is None:
        return {"tier": "unknown", "label": "Date unknown", "fresh": False, "days": None}
    # Say "Posted" only when we actually know the posting age; otherwise we're
    # reporting when we first found the role, so don't overstate it.
    verb = "Posted" if _posted_age_days(posted, None) is not None else "Found"
    if days <= 0:
        label = f"{verb} today"
    elif days == 1:
        label = f"{verb} yesterday"
    elif days <= 30:
        label = f"{verb} {days} days ago"
    else:
        label = f"{verb} 30+ days ago"
    if days <= _FRESH_MAX_DAYS:
        tier = "fresh"
    elif days <= 7:
        tier = "recent"
    else:
        tier = "old"
    return {"tier": tier, "label": label, "fresh": tier == "fresh", "days": days}


_STATUS_RANK = {"open": 0, "unknown": 1, "closed": 2}


def _with_probability(rows: list[dict]) -> list[dict]:
    """Annotate each row with selection probability + REAL open/closed status
    (verified by fetching the URL, cached), then sort: verified-open first,
    unknown next, verified-closed last; highest match within each group."""
    from utils.job_status import get_status
    for r in rows:
        r["prob"] = selection_probability(r.get("AI Score"))
        r["window"] = application_window(r.get("Posted"), r.get("Date Found"))
        r["fresh"] = freshness(r.get("Posted"), r.get("Date Found"))
        r["live"] = get_status(r.get("Job URL", ""))
    rows.sort(key=lambda r: (_STATUS_RANK.get(r["live"]["status"], 1), -r["prob"]["pct"]))
    return rows


# ── Platform / source grouping ───────────────────────────────────────────────
# Maps a stored Source value to a friendly platform label for the Manual page's
# per-platform tabs. Direct scrapers and Gmail-alert jobs both store a lowercase
# key here (linkedin / naukri / indeed / …), so grouping is uniform.
_SOURCE_LABELS = {
    "linkedin": "LinkedIn", "naukri": "Naukri", "indeed": "Indeed",
    "remotive": "Remotive", "remoteok": "RemoteOK", "rss": "RSS",
    "gmail": "Email", "email": "Email",
}


def source_key(row: dict) -> str:
    """Normalised lowercase source key for a job row ('' -> 'other')."""
    return str(row.get("Source") or "").strip().lower() or "other"


def source_label(key: str) -> str:
    """Friendly platform label for a source key."""
    return _SOURCE_LABELS.get(key, key.title())


# Job-board platforms always shown as their own space, even at 0 jobs, so the
# separation is visible before/while jobs arrive from each.
ALWAYS_SHOW = ("linkedin", "naukri", "indeed")


def source_tabs(rows: list[dict], always: tuple = ALWAYS_SHOW) -> list[dict]:
    """[{key,label,count}] per platform: every platform present in `rows`, plus
    the `always` job boards even when they have 0 jobs. Most jobs first."""
    counts: dict[str, int] = {}
    for r in rows:
        counts[source_key(r)] = counts.get(source_key(r), 0) + 1
    for k in always:
        counts.setdefault(k, 0)
    tabs = [{"key": k, "label": source_label(k), "count": n}
            for k, n in counts.items()]
    tabs.sort(key=lambda t: (-t["count"], t["label"]))
    return tabs


def _is_new(row: dict) -> bool:
    """True if this job was posted or first found today (fresh from a run)."""
    today = date.today()
    for field in ("Posted", "Date Found"):
        d = _parse_date(row.get(field))
        if d == today:
            return True
    return False


# ── Reads ─────────────────────────────────────────────────────────────────────
def manual_jobs(include_done: bool = False, new_only: bool = False,
                fresh_only: bool = False) -> list[dict]:
    """Jobs in 'Manual Apply Needed', ranked by selection probability.

    By default only those still Pending. Highest-chance roles come first so the
    UI can present the best bets at the top. `new_only` keeps just today's jobs;
    `fresh_only` keeps just recently-posted roles (within _FRESH_MAX_DAYS),
    sorted freshest-first so newly-posted jobs are easy to spot.
    """
    rows = _rows_as_dicts(S_MANUAL)
    if not include_done:
        rows = [r for r in rows if str(r.get("Status", "")).strip().lower() == "pending"]
    for r in rows:
        r["is_new"] = _is_new(r)
    rows = _with_probability(rows)
    if new_only:
        rows = [r for r in rows if r["is_new"]]
    if fresh_only:
        rows = [r for r in rows if r["fresh"]["fresh"]]
        rows.sort(key=lambda r: (r["fresh"]["days"] if r["fresh"]["days"] is not None else 999,
                                 -r["prob"]["pct"]))
    return rows


def walkins() -> list[dict]:
    return _rows_as_dicts(S_WALKIN)


def applied_jobs() -> list[dict]:
    """Applied jobs, each annotated with its selection probability."""
    rows = _rows_as_dicts(S_APPLIED)
    for r in rows:
        r["prob"] = selection_probability(r.get("AI Score"))
    return rows


def _date_label(d) -> str:
    """Human label for a day group: Today / Yesterday / N days ago / date."""
    if d is None:
        return "Date unknown"
    delta = (date.today() - d).days
    if delta <= 0:
        return "Today"
    if delta == 1:
        return "Yesterday"
    if delta < 7:
        return f"{delta} days ago"
    return d.strftime("%d %b %Y")


def applied_history() -> list[dict]:
    """Applied jobs grouped by the day you applied, newest day first.

    Returns a list of {key, date, label, count, jobs} groups so the History
    page can render a day-by-day timeline (Today, Yesterday, earlier dates).
    """
    groups: dict[str, dict] = {}
    for j in applied_jobs():
        d = _parse_date(j.get("Date Applied"))
        key = d.isoformat() if d else "unknown"
        groups.setdefault(key, {"key": key, "date": d, "jobs": []})["jobs"].append(j)

    dated = sorted((g for g in groups.values() if g["date"]),
                   key=lambda g: g["date"], reverse=True)
    undated = [g for g in groups.values() if not g["date"]]
    ordered = dated + undated
    for g in ordered:
        g["label"] = _date_label(g["date"])
        g["count"] = len(g["jobs"])
    return ordered


def pipeline_summary() -> dict:
    return get_pipeline_summary()


def manual_job_by_url(url: str) -> "dict | None":
    """Find a single Manual-sheet job (any status) by its URL."""
    if not url:
        return None
    return next((j for j in manual_jobs(include_done=True)
                 if j.get("Job URL") == url), None)


_RESPONDED = {"Shortlisted", "Interview Scheduled", "Interview Done", "Offer Received"}


def insights() -> dict:
    """Response-rate analytics over Applied Jobs: what's actually working.

    'Responded' = any status past the application stage (shortlist/interview/
    offer). Grouped by source and by match tier so you can double down on the
    channels that get replies.
    """
    jobs = applied_jobs()

    def _responded(j) -> bool:
        return str(j.get("Status", "")).strip() in _RESPONDED

    def _group(key_fn) -> list[dict]:
        buckets: dict[str, dict] = {}
        for j in jobs:
            k = key_fn(j)
            b = buckets.setdefault(k, {"name": k, "applied": 0, "responded": 0})
            b["applied"] += 1
            if _responded(j):
                b["responded"] += 1
        for b in buckets.values():
            b["rate"] = round(b["responded"] / b["applied"] * 100) if b["applied"] else 0
        return sorted(buckets.values(), key=lambda b: (-b["applied"], -b["rate"]))

    total = len(jobs)
    responded = sum(1 for j in jobs if _responded(j))
    return {
        "total": total,
        "responded": responded,
        "response_rate": round(responded / total * 100) if total else 0,
        "interviews": sum(1 for j in jobs if "Interview" in str(j.get("Status", ""))),
        "offers": sum(1 for j in jobs if "Offer" in str(j.get("Status", ""))),
        "by_source": _group(lambda j: j.get("Source") or "—"),
        "by_tier": _group(lambda j: j["prob"]["label"]),
    }


def pending_cold_mails() -> list[dict]:
    return cold_mailer.list_pending()


def counts() -> dict:
    """Headline numbers for the dashboard cards."""
    pending = manual_jobs()
    return {
        "manual_pending": len(pending),
        "manual_new": sum(1 for j in pending if j.get("is_new")),
        "manual_fresh": sum(1 for j in pending if j.get("fresh", {}).get("fresh")),
        "walkins": len(walkins()),
        "applied": len(applied_jobs()),
        "cold_pending": len(pending_cold_mails()),
    }


# ── Writes ──────────────────────────────────────────────────────────────────--
def _set_status_by_url(sheet: str, url_header: str, status_header: str,
                       url: str, new_status: str) -> bool:
    """Set the status cell of the row whose URL column matches `url`."""
    headers = HEADERS[sheet]
    url_col = headers.index(url_header) + 1
    status_col = headers.index(status_header) + 1
    with WB_LOCK:
        try:
            wb = load_workbook(EXCEL_FILE)
        except PermissionError as exc:
            raise WorkbookLocked("Close job_applications.xlsx in Excel and try again.") from exc
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=url_col).value == url:
                ws.cell(row=r, column=status_col, value=new_status)
                save_workbook(wb)
                return True
        return False


def mark_manual_applied(url: str) -> bool:
    """Move a 'Manual Apply Needed' job into 'Applied Jobs' and mark it done."""
    if not url:
        return False
    job = next((j for j in manual_jobs(include_done=True)
                if j.get("Job URL") == url), None)
    if job is None:
        logger.warning("mark_manual_applied: url not found %s", url)
        return False
    log_applied({
        "company": job.get("Company Name", ""),
        "title": job.get("Job Title", ""),
        "location": job.get("Location", ""),
        "source": job.get("Source", ""),
        "url": url,
        "notes": "Marked applied from web UI",
    })
    _set_status_by_url(S_MANUAL, "Job URL", "Status", url, "Applied")
    refresh_dashboard()
    return True


def dismiss_manual(url: str) -> bool:
    """Drop a manual job off the to-apply list by marking it Dismissed."""
    if not url:
        return False
    return _set_status_by_url(S_MANUAL, "Job URL", "Status", url, "Dismissed")


def clear_closed_manual() -> int:
    """Mark every still-pending, likely-closed manual job as Dismissed.

    Uses REAL verified status (run Verify first); clears out postings confirmed
    to be closed so the dashboard shows current roles only. Returns the count.
    """
    closed_urls = {j.get("Job URL") for j in manual_jobs()
                   if j["live"]["status"] == "closed"}
    if not closed_urls:
        return 0
    headers = HEADERS[S_MANUAL]
    url_col = headers.index("Job URL") + 1
    status_col = headers.index("Status") + 1
    with WB_LOCK:
        try:
            wb = load_workbook(EXCEL_FILE)
        except PermissionError as exc:
            raise WorkbookLocked("Close job_applications.xlsx in Excel and try again.") from exc
        ws = wb[S_MANUAL]
        cleared = 0
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=status_col).value).strip().lower() != "pending":
                continue
            if ws.cell(row=r, column=url_col).value in closed_urls:
                ws.cell(row=r, column=status_col, value="Dismissed")
                cleared += 1
        if cleared:
            save_workbook(wb)
        wb.close()
    logger.info("Cleared %d likely-closed manual job(s).", cleared)
    return cleared


def restore_dismissed_manual() -> int:
    """Bring dismissed/cleared manual jobs back to Pending so they show again.

    Useful when 'Clear all' was hit but the inbox has no fresher jobs to replace
    them — re-fetching can't re-add them (they dedupe), so we un-dismiss instead.
    Returns how many were restored.
    """
    headers = HEADERS[S_MANUAL]
    status_col = headers.index("Status") + 1
    with WB_LOCK:
        try:
            wb = load_workbook(EXCEL_FILE)
        except PermissionError as exc:
            raise WorkbookLocked("Close job_applications.xlsx in Excel and try again.") from exc
        ws = wb[S_MANUAL]
        restored = 0
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=status_col).value).strip().lower() == "dismissed":
                ws.cell(row=r, column=status_col, value="Pending")
                restored += 1
        if restored:
            save_workbook(wb)
        wb.close()
    logger.info("Restored %d dismissed manual job(s).", restored)
    return restored


def clear_all_manual() -> int:
    """Dismiss every still-pending manual job (wipe the whole to-apply list).

    Dismiss (not hard-delete) keeps the rows for dedup so they don't re-appear
    on the next fetch. Returns how many were cleared.
    """
    headers = HEADERS[S_MANUAL]
    status_col = headers.index("Status") + 1
    with WB_LOCK:
        try:
            wb = load_workbook(EXCEL_FILE)
        except PermissionError as exc:
            raise WorkbookLocked("Close job_applications.xlsx in Excel and try again.") from exc
        ws = wb[S_MANUAL]
        cleared = 0
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=status_col).value).strip().lower() == "pending":
                ws.cell(row=r, column=status_col, value="Dismissed")
                cleared += 1
        if cleared:
            save_workbook(wb)
        wb.close()
    logger.info("Cleared ALL %d pending manual job(s).", cleared)
    return cleared


_DELETE_URL_HEADER = {S_MANUAL: "Job URL", S_WALKIN: "Job URL", S_APPLIED: "Job URL"}


def delete_by_url(sheet: str, url: str) -> bool:
    """Permanently remove the row whose Job URL matches `url` from `sheet`."""
    if not url or sheet not in _DELETE_URL_HEADER:
        return False
    col = HEADERS[sheet].index(_DELETE_URL_HEADER[sheet]) + 1
    with WB_LOCK:
        try:
            wb = load_workbook(EXCEL_FILE)
        except PermissionError as exc:
            raise WorkbookLocked("Close job_applications.xlsx in Excel and try again.") from exc
        ws = wb[sheet]
        deleted = False
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=col).value == url:
                ws.delete_rows(r, 1)
                deleted = True
                break
        if deleted:
            save_workbook(wb)
        wb.close()
    if deleted and sheet == S_APPLIED:
        refresh_dashboard()
    return deleted


def discard_cold_mail(email: str, company: str = "") -> bool:
    """Remove a queued (pending-approval) cold mail draft."""
    return cold_mailer.discard_pending(email, company)


def verify_manual_status() -> dict:
    """Fetch each pending manual job's URL to record its REAL open/closed status.

    Updates the on-disk status cache; the UI then shows verified status and
    sinks closed roles. Returns a small summary for the flash message.
    """
    from collections import Counter

    from utils.job_status import verify_urls
    urls = [j.get("Job URL") for j in manual_jobs() if j.get("Job URL")]
    results = verify_urls(urls)
    c = Counter(v["status"] for v in results.values())
    return {"checked": len(results), "open": c.get("open", 0),
            "closed": c.get("closed", 0), "unknown": c.get("unknown", 0)}


def set_applied_status(url: str, new_status: str) -> bool:
    """Update the Status of a row in 'Applied Jobs' (reuses status_manager)."""
    try:
        ok = update_status(url, new_status)
    except PermissionError as exc:
        raise WorkbookLocked("Close job_applications.xlsx in Excel and try again.") from exc
    if ok:
        refresh_dashboard()
    return ok


def approve_cold_mails(limit: int | None = None) -> dict:
    return cold_mailer.send_approved(limit=limit)


def backfill_manual_scores() -> int:
    """Give legacy Manual rows a real fit score (and probability).

    Rows logged before scores were persisted read back as 0 -> a flat 4% for
    everything. Here we re-score each such row from its title/company (the only
    text those rows kept), so the percentage actually differentiates roles.
    New pipeline runs already store the full title+description score. Returns
    the number of rows updated.
    """
    from ai.job_scorer import score_job
    from tracker.excel_tracker import HEADERS, S_MANUAL

    init_excel()
    headers = HEADERS[S_MANUAL]
    col = {h: headers.index(h) + 1 for h in ("Job Title", "Company Name", "AI Score")}
    with WB_LOCK:
        try:
            wb = load_workbook(EXCEL_FILE)
        except PermissionError as exc:
            raise WorkbookLocked("Close job_applications.xlsx in Excel and try again.") from exc
        ws = wb[S_MANUAL]
        updated = 0
        for r in range(2, ws.max_row + 1):
            title = ws.cell(row=r, column=col["Job Title"]).value
            if not title:
                continue
            if ws.cell(row=r, column=col["AI Score"]).value not in (None, "", 0):
                continue  # already scored
            company = ws.cell(row=r, column=col["Company Name"]).value or ""
            score = score_job("", str(title), str(company))
            ws.cell(row=r, column=col["AI Score"], value=score)
            updated += 1
        if updated:
            save_workbook(wb)
        wb.close()
    logger.info("Backfilled AI Score on %d legacy Manual rows.", updated)
    return updated
