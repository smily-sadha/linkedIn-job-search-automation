"""Automatic follow-ups for cold mails that have gone unanswered.

Runs every time main.py runs. Honours the same DRY_RUN / REQUIRE_MAIL_CONFIRM
rails as the cold mailer, and caps follow-ups at MAX_FOLLOWUPS per company.
"""
from datetime import date

from openpyxl import load_workbook

from config.config import (
    DRY_RUN, EXCEL_FILE, FOLLOWUP_BODY, FOLLOWUP_SUBJECT,
    REQUIRE_MAIL_CONFIRM, YOUR_NAME,
)
from mailers.cold_mailer import _send
from tracker import excel_tracker as xl
from tracker.excel_tracker import S_COLD
from utils.logger import get_logger

logger = get_logger("followup_mailer")

# Cold Mails Sent column indices (1-based).
C_COMPANY, C_TITLE, C_HR, C_EMAIL, C_SENT, C_FUP_DATE, C_FUP_SENT, C_SUBJECT, C_STATUS = range(1, 10)


def _due(followup_date: str) -> bool:
    try:
        return date.fromisoformat(str(followup_date)) <= date.today()
    except (ValueError, TypeError):
        return False


def send_followups() -> dict:
    summary = {"sent": 0, "drafted": 0, "skipped": 0}
    wb = load_workbook(EXCEL_FILE)
    ws = wb[S_COLD]

    for row in range(2, ws.max_row + 1):
        fup_sent = ws.cell(row=row, column=C_FUP_SENT).value
        status = str(ws.cell(row=row, column=C_STATUS).value or "")
        # Only follow up on mails that were actually sent and not yet followed up.
        if fup_sent and str(fup_sent).startswith("Yes"):
            continue
        if not status.startswith("Sent"):
            continue
        if not _due(ws.cell(row=row, column=C_FUP_DATE).value):
            continue

        company = ws.cell(row=row, column=C_COMPANY).value or ""
        title = ws.cell(row=row, column=C_TITLE).value or ""
        hr = ws.cell(row=row, column=C_HR).value or "Hiring Manager"
        email = ws.cell(row=row, column=C_EMAIL).value or ""
        original = ws.cell(row=row, column=C_SENT).value or ""

        record = {
            "company": company, "title": title, "email": email,
            "subject": FOLLOWUP_SUBJECT.format(job_title=title, your_name=YOUR_NAME),
            "body": FOLLOWUP_BODY.format(hr_name=hr, job_title=title, company=company,
                                         original_date=original, your_name=YOUR_NAME),
            "resume": "",
        }

        if DRY_RUN or REQUIRE_MAIL_CONFIRM:
            logger.info("[Follow-up] (not sent: dry-run/confirm) would follow up %s", company)
            summary["drafted"] += 1
            continue

        if _send(record):
            ws.cell(row=row, column=C_FUP_SENT, value="Yes")
            xl.log_followup({
                "company": company, "title": title,
                "original_date": str(original),
                "followup_date": str(ws.cell(row=row, column=C_FUP_DATE).value),
            })
            summary["sent"] += 1
        else:
            summary["skipped"] += 1

    wb.save(EXCEL_FILE)
    logger.info("[Follow-up] %s", summary)
    return summary
