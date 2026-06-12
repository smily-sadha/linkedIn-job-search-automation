"""
Telegram remote control for the job-hunt assistant.

Run with:  python -m telegram_bot.bot

Security: only messages from TELEGRAM_CHAT_ID are honoured; everything else is
ignored. Long-running scans run in a worker thread so the bot stays responsive.

Commands:
  /start | /help   show this help
  /run [source]    run all sources now (or one: remotive | remoteok | rss)
  /status          pipeline summary from the tracker
  /walkins         upcoming walk-in drives
  /manual          pending "manual apply" jobs (max 10)
  /followups       cold mails awaiting a follow-up
  /skipped         last 10 skipped jobs with reasons
  /approve [N]     send up to N queued cold mails (all if omitted)
  /stop            stop this bot process
"""
import asyncio
from datetime import datetime

from openpyxl import load_workbook
from telegram import Update
from telegram.ext import (
    Application, CommandHandler, ContextTypes, MessageHandler, filters,
)

from config.config import EXCEL_FILE, TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID
from mailers import cold_mailer
from sources.registry import available_keys
from tracker.excel_tracker import (
    S_FOLLOWUP, S_MANUAL, S_SKIPPED, S_WALKIN, init_excel,
)
from tracker.status_manager import get_pipeline_summary
from utils.logger import get_logger

logger = get_logger("telegram_bot")

HELP = (
    "Job-Hunt Assistant\n"
    "/run [source]  - run now (source: " + " | ".join(available_keys()) + ")\n"
    "/status        - pipeline summary\n"
    "/walkins       - upcoming walk-in drives\n"
    "/manual        - pending manual-apply jobs\n"
    "/followups     - cold mails awaiting follow-up\n"
    "/skipped       - last 10 skipped jobs\n"
    "/approve [N]   - send queued cold mails\n"
    "/stop          - stop the bot\n"
    "/help          - this message"
)


def _authorised(update: Update) -> bool:
    return str(update.effective_chat.id) == str(TELEGRAM_CHAT_ID)


def _guard(func):
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not _authorised(update):
            logger.warning("Ignored message from unauthorised chat %s", update.effective_chat.id)
            return
        await func(update, context)
    return wrapper


def _read_rows(sheet: str, limit: int | None = None) -> list[tuple]:
    init_excel()
    wb = load_workbook(EXCEL_FILE, read_only=True)
    ws = wb[sheet]
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(r)]
    wb.close()
    return rows[:limit] if limit else rows


# ── command handlers ─────────────────────────────────────────────────────────
@_guard
async def cmd_start(update, context):
    await update.message.reply_text(HELP)


@_guard
async def cmd_run(update, context):
    source = (context.args[0].lower() if context.args else "all")
    await update.message.reply_text(f"Starting run (source: {source}) ...")

    # run_all is blocking/synchronous -> push to a worker thread.
    from main import run_all
    loop = asyncio.get_running_loop()
    summary = await loop.run_in_executor(None, run_all, source)

    await update.message.reply_text(
        f"Run complete.\n"
        f"Found: {summary['found']}  Walk-ins: {summary['walkins']}\n"
        f"Cold mails: {summary['cold_mails']}  Manual: {summary['queued_manual']}\n"
        f"Skipped: {summary['skipped']}  Errors: {summary['errors']}\n"
        f"Duration: {summary['duration_seconds']}s"
    )


@_guard
async def cmd_status(update, context):
    summary = get_pipeline_summary()
    pending = len(cold_mailer.list_pending())
    lines = [f"Job Hunt Status - {datetime.now():%d %b %Y}"]
    lines += [f"{k}: {v}" for k, v in summary.items()]
    lines.append(f"Cold mails awaiting approval: {pending}")
    await update.message.reply_text("\n".join(lines))


@_guard
async def cmd_walkins(update, context):
    rows = _read_rows(S_WALKIN)
    if not rows:
        await update.message.reply_text("No walk-in drives logged yet.")
        return
    msg = "Walk-In Drives:\n" + "\n".join(
        f"- {r[0]} | {r[1]} | {r[3]} {r[4]} @ {r[5]}" for r in rows[:15])
    await update.message.reply_text(msg)


@_guard
async def cmd_manual(update, context):
    rows = _read_rows(S_MANUAL, limit=10)
    if not rows:
        await update.message.reply_text("No pending manual-apply jobs.")
        return
    msg = "Manual Apply Needed:\n" + "\n".join(
        f"- {r[0]} | {r[1]}\n  {r[5]}" for r in rows)
    await update.message.reply_text(msg)


@_guard
async def cmd_followups(update, context):
    rows = _read_rows(S_FOLLOWUP)
    if not rows:
        await update.message.reply_text("No follow-ups logged yet.")
        return
    msg = "Follow-Ups:\n" + "\n".join(f"- {r[0]} | {r[1]} | sent {r[4]}" for r in rows[:15])
    await update.message.reply_text(msg)


@_guard
async def cmd_skipped(update, context):
    rows = _read_rows(S_SKIPPED)
    if not rows:
        await update.message.reply_text("No skipped jobs logged yet.")
        return
    msg = "Last skipped jobs:\n" + "\n".join(
        f"- {r[0]} | {r[1]} ({r[4]})" for r in rows[-10:])
    await update.message.reply_text(msg)


@_guard
async def cmd_approve(update, context):
    limit = int(context.args[0]) if context.args and context.args[0].isdigit() else None
    loop = asyncio.get_running_loop()
    result = await loop.run_in_executor(None, lambda: cold_mailer.send_approved(limit=limit))
    await update.message.reply_text(
        f"Sent: {result['sent']}  Failed: {result['failed']}  Remaining: {result['remaining']}")


@_guard
async def cmd_stop(update, context):
    await update.message.reply_text("Stopping bot. Bye!")
    context.application.stop_running()


@_guard
async def unknown(update, context):
    await update.message.reply_text("Unknown command. Try /help")


def build_app() -> Application:
    if not TELEGRAM_BOT_TOKEN:
        raise RuntimeError("TELEGRAM_BOT_TOKEN is not set (see .env).")
    app = Application.builder().token(TELEGRAM_BOT_TOKEN).build()
    app.add_handler(CommandHandler(["start", "help"], cmd_start))
    app.add_handler(CommandHandler("run", cmd_run))
    app.add_handler(CommandHandler("status", cmd_status))
    app.add_handler(CommandHandler("walkins", cmd_walkins))
    app.add_handler(CommandHandler("manual", cmd_manual))
    app.add_handler(CommandHandler("followups", cmd_followups))
    app.add_handler(CommandHandler("skipped", cmd_skipped))
    app.add_handler(CommandHandler("approve", cmd_approve))
    app.add_handler(CommandHandler("stop", cmd_stop))
    app.add_handler(MessageHandler(filters.COMMAND, unknown))
    return app


def main() -> None:
    logger.info("Starting Telegram bot ...")
    app = build_app()
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
