"""Update and summarise application pipeline status in the Applied Jobs sheet."""
from openpyxl import load_workbook

from config.config import EXCEL_FILE
from tracker.excel_tracker import S_APPLIED, VALID_STATUSES
from utils.logger import get_logger

logger = get_logger("status_manager")

_URL_COL = 8     # H
_STATUS_COL = 9  # I


def update_status(job_url: str, new_status: str) -> bool:
    """Find the row by Job URL and set its Status. Returns True on success."""
    if new_status not in VALID_STATUSES:
        logger.warning("Invalid status '%s' (allowed: %s)", new_status, VALID_STATUSES)
        return False
    wb = load_workbook(EXCEL_FILE)
    ws = wb[S_APPLIED]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=_URL_COL).value == job_url:
            ws.cell(row=row, column=_STATUS_COL, value=new_status)
            wb.save(EXCEL_FILE)
            logger.info("Status updated: %s -> %s", job_url, new_status)
            return True
    logger.warning("Job URL not found: %s", job_url)
    return False


def get_pipeline_summary() -> dict:
    """Return a count of each status across all applied jobs."""
    wb = load_workbook(EXCEL_FILE, read_only=True)
    ws = wb[S_APPLIED]
    summary = {s: 0 for s in VALID_STATUSES}
    for row in ws.iter_rows(min_row=2, min_col=_STATUS_COL, max_col=_STATUS_COL, values_only=True):
        status = row[0]
        if status in summary:
            summary[status] += 1
    wb.close()
    return summary
