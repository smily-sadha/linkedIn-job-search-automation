"""Color-coded Excel tracker with 7 sheets, built on openpyxl.

Public helpers used by the rest of the app:
    init_excel()                  - create the workbook if missing
    is_duplicate(url)             - True if URL already logged in Applied/Walk-in/Cold/Manual
    log_applied(job)              - Sheet 2
    log_walkin(job)               - Sheet 3
    log_manual(job, reason)       - Sheet 4
    log_cold_mail(record)         - Sheet 5
    log_followup(record)          - Sheet 6
    log_skipped(job, reason)      - Sheet 7
    refresh_dashboard()           - recompute Sheet 1 summary
"""
import os
import tempfile
import threading
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config.config import EXCEL_FILE
from utils.logger import get_logger

logger = get_logger("excel_tracker")

# Background pipeline runs and web requests share this process, so all workbook
# open/save must be serialised — otherwise a read can hit a file that's mid-save
# (openpyxl rewrites the whole zip on save) and blow up with BadZipFile.
# Re-entrant so nested helpers (e.g. log_applied -> _append) don't deadlock.
WB_LOCK = threading.RLock()
_MIGRATED = False  # run the schema migration once per process, not per read


def save_workbook(wb) -> None:
    """Atomically save `wb` to EXCEL_FILE.

    Writes to a temp file in the same folder, then os.replace()s it into place —
    an atomic operation, so a crash mid-save (or a concurrent read) can never
    leave a half-written / corrupt workbook. This is what stops the BadZipFile
    corruption that previously truncated data. Call inside WB_LOCK.
    """
    target = str(EXCEL_FILE)
    folder = os.path.dirname(target) or "."
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=folder)
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, target)  # atomic on the same filesystem
    finally:
        if os.path.exists(tmp):
            try:
                os.remove(tmp)
            except OSError:
                pass

# ── Sheet names ────────────────────────────────────────────────────────────
S_DASH = "Dashboard"
S_APPLIED = "Applied Jobs"
S_WALKIN = "Walk-In Drives"
S_MANUAL = "Manual Apply Needed"
S_COLD = "Cold Mails Sent"
S_FOLLOWUP = "Follow-Ups"
S_SKIPPED = "Skipped Jobs"

HEADERS = {
    S_APPLIED: ["Company Name", "Job Title", "Location", "Date Applied", "Source",
                "AI Score", "Resume Used", "Job URL", "Status", "Notes"],
    S_WALKIN: ["Company Name", "Job Title", "Location", "Walk-In Date", "Walk-In Time",
               "Venue", "Source", "Job URL", "Contact", "Status"],
    S_MANUAL: ["Company Name", "Job Title", "Location", "Date Found", "Source",
               "Job URL", "Reason", "Status", "AI Score", "Posted", "Match Reason"],
    S_COLD: ["Company Name", "Job Title", "HR Name", "Recipient Email", "Date Sent",
             "Follow-Up Date", "Follow-Up Sent", "Subject", "Status"],
    S_FOLLOWUP: ["Company Name", "Job Title", "Original Date", "Follow-Up Date",
                 "Sent Date", "Status"],
    S_SKIPPED: ["Company Name", "Job Title", "Source", "Date", "Reason"],
}

HEADER_FILLS = {
    S_APPLIED: "1F4E79",
    S_WALKIN: "375623",
    S_MANUAL: "843C0C",
    S_COLD: "4A235A",
    S_FOLLOWUP: "1A5276",
    S_SKIPPED: "424242",
}

VALID_STATUSES = [
    "Applied", "Viewed", "Shortlisted", "Interview Scheduled",
    "Interview Done", "Offer Received", "Rejected", "Ghosted",
]

_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_CELL_FONT = Font(name="Arial", size=10)
_ALT_FILL = PatternFill("solid", fgColor="EBF5FB")


def _style_header(ws, sheet_name: str) -> None:
    fill = PatternFill("solid", fgColor=HEADER_FILLS[sheet_name])
    for col, title in enumerate(HEADERS[sheet_name], start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = _HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(title) + 4)
    ws.freeze_panes = "A2"


def init_excel() -> None:
    """Create the workbook with all 7 sheets if it does not yet exist."""
    path = Path(EXCEL_FILE)
    if path.exists():
        _migrate_schema()
        return
    with WB_LOCK:
        if path.exists():  # another thread may have created it while we waited
            _migrate_schema()
            return
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        # Dashboard is the default first sheet.
        dash = wb.active
        dash.title = S_DASH

        for name in (S_APPLIED, S_WALKIN, S_MANUAL, S_COLD, S_FOLLOWUP, S_SKIPPED):
            ws = wb.create_sheet(name)
            _style_header(ws, name)

        # Status dropdown on Applied Jobs.
        applied = wb[S_APPLIED]
        dv = DataValidation(type="list", formula1='"%s"' % ",".join(VALID_STATUSES), allow_blank=True)
        applied.add_data_validation(dv)
        dv.add("I2:I1000")

        _build_dashboard_skeleton(dash)
        save_workbook(wb)
        logger.info("Created Excel tracker with 7 sheets: %s", EXCEL_FILE)


def _build_dashboard_skeleton(ws) -> None:
    ws["A1"] = "Job Hunt Dashboard"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="1F4E79")
    ws.merge_cells("A1:C1")
    labels = [
        ("Total Applied", 3),
        ("Walk-Ins Found", 4),
        ("Cold Mails Sent", 5),
        ("Manual Pending", 6),
        ("Shortlisted", 7),
        ("Interviews", 8),
        ("Offers", 9),
        ("Response Rate", 10),
        ("Best Source", 11),
        ("Last Updated", 12),
    ]
    for label, row in labels:
        ws.cell(row=row, column=1, value=label).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=row, column=2, value=0).font = _CELL_FONT
    ws["B11"] = "-"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 24


def _migrate_schema() -> None:
    """Backfill header columns added after a workbook was first created.

    New columns ('AI Score', 'Posted') are always appended to the end of the
    Manual sheet, so every existing column stays aligned and legacy rows simply
    read back as blank for the new fields. Runs at most once per process (reads
    happen constantly, but the schema only needs checking once). Idempotent; no
    save when already current or when the workbook is open in Excel.
    """
    global _MIGRATED
    if _MIGRATED:
        return
    with WB_LOCK:
        if _MIGRATED:
            return
        try:
            wb = load_workbook(EXCEL_FILE)
        except PermissionError:
            return  # open in Excel — retried on the next run
        ws = wb[S_MANUAL]
        changed = False
        for idx, header in enumerate(HEADERS[S_MANUAL], start=1):
            if ws.cell(row=1, column=idx).value != header:
                cell = ws.cell(row=1, column=idx, value=header)
                cell.font = _HEADER_FONT
                cell.fill = PatternFill("solid", fgColor=HEADER_FILLS[S_MANUAL])
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = _BORDER
                ws.column_dimensions[get_column_letter(idx)].width = max(12, len(header) + 4)
                changed = True
        if changed:
            save_workbook(wb)
            logger.info("Migrated Manual sheet headers to: %s", HEADERS[S_MANUAL])
        wb.close()
        _MIGRATED = True


# ── Low-level append helper ──────────────────────────────────────────────────
def _append(sheet_name: str, row_values: list) -> None:
    with WB_LOCK:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]
        r = ws.max_row + 1
        for col, val in enumerate(row_values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _CELL_FONT
            cell.border = _BORDER
            if r % 2 == 0:
                cell.fill = _ALT_FILL
        # AI-score conditional color on Applied Jobs (column 6).
        if sheet_name == S_APPLIED and isinstance(row_values[5], int):
            score = row_values[5]
            color = "C6EFCE" if score >= 8 else ("FFEB9C" if score >= 6 else "FFC7CE")
            ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor=color)
        save_workbook(wb)


def _all_urls() -> set[str]:
    """Collect every Job URL already recorded so we can dedupe."""
    with WB_LOCK:
        wb = load_workbook(EXCEL_FILE, read_only=True)
        urls = set()
        url_col = {S_APPLIED: 8, S_WALKIN: 8, S_MANUAL: 6}
        for sheet, col in url_col.items():
            if sheet in wb.sheetnames:
                for row in wb[sheet].iter_rows(min_row=2, min_col=col, max_col=col, values_only=True):
                    if row[0]:
                        urls.add(str(row[0]))
        wb.close()
        return urls


def is_duplicate(url: str) -> bool:
    if not url:
        return False
    return url in _all_urls()


# ── Public log_* helpers ─────────────────────────────────────────────────────
def log_applied(job: dict) -> None:
    _append(S_APPLIED, [
        job.get("company", ""), job.get("title", ""), job.get("location", ""),
        date.today().isoformat(), job.get("source", ""), job.get("score", 0),
        Path(job.get("resume", "")).name, job.get("url", ""), "Applied",
        job.get("notes", ""),
    ])


def log_walkin(job: dict) -> None:
    _append(S_WALKIN, [
        job.get("company", ""), job.get("title", ""), job.get("location", ""),
        job.get("walkin_date", ""), job.get("walkin_time", ""), job.get("venue", ""),
        job.get("source", ""), job.get("url", ""), job.get("contact", ""), "To Attend",
    ])


def log_manual(job: dict, reason: str) -> None:
    _append(S_MANUAL, [
        job.get("company", ""), job.get("title", ""), job.get("location", ""),
        date.today().isoformat(), job.get("source", ""), job.get("url", ""),
        reason, "Pending", job.get("score", 0), job.get("posted", ""),
        job.get("match_reason", ""),
    ])


def log_cold_mail(record: dict) -> None:
    sent = date.today()
    _append(S_COLD, [
        record.get("company", ""), record.get("title", ""), record.get("hr_name", ""),
        record.get("email", ""), sent.isoformat(),
        (sent + timedelta(days=record.get("followup_days", 5))).isoformat(),
        "No", record.get("subject", ""), record.get("status", "Sent"),
    ])


def log_followup(record: dict) -> None:
    _append(S_FOLLOWUP, [
        record.get("company", ""), record.get("title", ""),
        record.get("original_date", ""), record.get("followup_date", ""),
        date.today().isoformat(), "Sent",
    ])


def log_skipped(job: dict, reason: str) -> None:
    _append(S_SKIPPED, [
        job.get("company", ""), job.get("title", ""), job.get("source", ""),
        date.today().isoformat(), reason,
    ])


# ── Dashboard refresh ────────────────────────────────────────────────────────
def _count_rows(ws) -> int:
    return max(0, ws.max_row - 1)


def refresh_dashboard() -> dict:
    """Recompute the Dashboard summary and return it as a dict."""
    with WB_LOCK:
        wb = load_workbook(EXCEL_FILE)
        applied = wb[S_APPLIED]

        total_applied = _count_rows(applied)
        statuses = [r[0] for r in applied.iter_rows(min_row=2, min_col=9, max_col=9, values_only=True) if r[0]]
        shortlisted = sum(1 for s in statuses if "Shortlisted" in str(s))
        interviews = sum(1 for s in statuses if "Interview" in str(s))
        offers = sum(1 for s in statuses if "Offer" in str(s))

        # Best source by applied count.
        sources = [r[0] for r in applied.iter_rows(min_row=2, min_col=5, max_col=5, values_only=True) if r[0]]
        best_source = max(set(sources), key=sources.count) if sources else "-"

        walkins = _count_rows(wb[S_WALKIN])
        cold = _count_rows(wb[S_COLD])
        manual = _count_rows(wb[S_MANUAL])
        response_rate = (shortlisted + interviews + offers) / total_applied * 100 if total_applied else 0

        values = {
            "Total Applied": total_applied, "Walk-Ins Found": walkins,
            "Cold Mails Sent": cold, "Manual Pending": manual,
            "Shortlisted": shortlisted, "Interviews": interviews, "Offers": offers,
            "Response Rate": f"{response_rate:.1f}%", "Best Source": best_source,
            "Last Updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        }

        dash = wb[S_DASH]
        row_map = {dash.cell(row=r, column=1).value: r for r in range(1, dash.max_row + 1)}
        for label, val in values.items():
            if label in row_map:
                dash.cell(row=row_map[label], column=2, value=val)
        dash.cell(row=row_map["Best Source"], column=2).fill = PatternFill("solid", fgColor="FFF2CC")
        save_workbook(wb)
        return values
